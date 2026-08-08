import asyncio
import base64
import hashlib
import hmac
import io
import json
import os
import re
import subprocess
import sys
import time
import urllib.parse
from datetime import datetime, timedelta, timezone

import cloudinary
import cloudinary.api
import cloudinary.uploader
from aiohttp import ClientSession, ClientTimeout, CookieJar, web
from playwright.async_api import async_playwright
from yarl import URL

# ================== CONFIG ==================
BOT_TOKEN = os.getenv("BOT_TOKEN")
ALLOWED_ADMINS = {5348697217, 547004364}
DATA_PATH = "/data/drivers.json"

# Optional driver fields (beyond the required name/car_model/car_number/tariff).
# Stored as trimmed strings; shared by add/update/get so the set stays in sync.
# NOTE: "name" is intentionally NOT updatable — it is part of the Cloudinary
# folder path (drivers/{id}_{name}); renaming would orphan stored documents.
OPTIONAL_DRIVER_FIELDS = [
    "car_year",
    "car_trim",
    "planet_gps_device_id",
    "monthly_mileage_limit",
    "weekly_price",
    "dmv_inspection_date",
    "payment_weekday",
    "last_service_mileage",
    "last_service_date",
]

PLANET_GPS_BASE = "https://web.planetgps.com"
PLANET_GPS_EMAIL = os.getenv("PLANET_GPS_EMAIL", "alexyss.waterry@icloud.com")
PLANET_GPS_PASSWORD = os.getenv("PLANET_GPS_PASSWORD", "")
PLANET_GPS_USER_ID = 272967

GPS_HTTP_TIMEOUT = ClientTimeout(total=30)
GPS_REPORT_TIMEOUT = ClientTimeout(total=180)  # Excel export can take minutes to build
GPS_PING_INTERVAL = 600          # keepalive ping every 10 min (ASP.NET sessions are sliding)
GPS_LOGIN_MAX_BACKOFF = 1800     # cap on the delay between failed login attempts

if not BOT_TOKEN:
    print("WARNING: BOT_TOKEN is not set — Telegram auth will reject everyone")
if not PLANET_GPS_PASSWORD:
    print("WARNING: PLANET_GPS_PASSWORD is not set — GPS login will fail")

cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET")
)


# ================== AUTH ==================
def verify_telegram_init_data(init_data: str) -> dict | None:
    try:
        parsed = dict(urllib.parse.parse_qsl(init_data, strict_parsing=True))
        received_hash = parsed.pop("hash", None)
        if not received_hash:
            return None
        data_check_string = "\n".join(f"{k}={v}" for k, v in sorted(parsed.items()))
        secret_key = hmac.new(b"WebAppData", BOT_TOKEN.encode(), hashlib.sha256).digest()
        expected_hash = hmac.new(secret_key, data_check_string.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(expected_hash, received_hash):
            return None
        auth_date = int(parsed.get("auth_date", 0))
        if time.time() - auth_date > 86400:
            return None
        user = json.loads(parsed.get("user", "{}"))
        return user
    except Exception:
        return None


def get_user_from_request(request: web.Request) -> dict | None:
    init_data = request.headers.get("X-Telegram-Init-Data", "")
    if not init_data:
        return None
    return verify_telegram_init_data(init_data)


def is_admin_request(request: web.Request) -> bool:
    user = get_user_from_request(request)
    return bool(user and user.get("id") in ALLOWED_ADMINS)


# ================== HELPERS ==================
def load_drivers() -> dict:
    if not os.path.exists(DATA_PATH):
        return {}
    try:
        with open(DATA_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        print(f"drivers.json read error: {e}")
        return {}


def save_drivers(data: dict):
    os.makedirs(os.path.dirname(DATA_PATH), exist_ok=True)
    tmp = DATA_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)
    os.replace(tmp, DATA_PATH)


def get_driver_folder(driver_id: str, driver_name: str) -> str:
    name = driver_name.replace(" ", "_")
    return f"drivers/{driver_id}_{name}"


def list_driver_files(folder: str) -> list[dict]:
    try:
        result = cloudinary.api.resources(type="upload", prefix=folder + "/", max_results=100)
        files = []
        for r in result.get("resources", []):
            files.append({
                "public_id": r["public_id"],
                "name": r["public_id"].split("/")[-1],
                "url": r["secure_url"],
                "created_at": r.get("created_at", ""),
                "format": r.get("format", ""),
                "bytes": r.get("bytes", 0),
            })
        return files
    except Exception:
        return []


# ================== PLANET GPS: SESSION CORE ==================
_gps_session: ClientSession | None = None
_gps_cookies: dict | None = None
_gps_lock = asyncio.Lock()
_login_fail_streak = 0
_next_login_allowed = 0.0  # time.monotonic() deadline

FLEET_PAYLOAD = json.dumps({
    "UserID": PLANET_GPS_USER_ID, "isFirst": True, "TimeZones": "5:00", "DeviceID": 0
})

GPS_BASE_HEADERS = {
    "Content-Type": "application/json",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Origin": PLANET_GPS_BASE,
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "X-Requested-With": "XMLHttpRequest",
}


def _kill_stray_browsers():
    """Safety net: reap zombie headless-chrome processes from failed launches."""
    if not sys.platform.startswith("linux"):
        return
    try:
        subprocess.run(["pkill", "-f", "chrome-headless-shell"],
                       capture_output=True, timeout=10)
    except Exception:
        pass


async def planet_gps_login_playwright() -> bool:
    """Log in via headless browser and stash the session cookies.

    The browser AND the playwright driver process are always closed, even on
    failure — leaking them exhausts the container's threads/PIDs and
    eventually makes every launch fail with pthread_create EAGAIN.
    """
    global _gps_cookies
    pw = None
    browser = None
    try:
        print("Starting Playwright login...")
        await asyncio.to_thread(_kill_stray_browsers)

        pw = await async_playwright().start()
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-zygote", "--disable-gpu"],
        )
        context = await browser.new_context()
        page = await context.new_page()

        await page.goto(f"{PLANET_GPS_BASE}/index.aspx", wait_until="domcontentloaded")
        await page.wait_for_timeout(2000)

        frame = page.frame_locator("iframe#ifm")
        await frame.locator("#changBar0").click()
        await page.wait_for_timeout(500)
        await frame.locator('input[name="txtUserName"]').fill(PLANET_GPS_EMAIL)
        await frame.locator('input[name="txtAccountPassword"]').fill(PLANET_GPS_PASSWORD)

        async with page.expect_navigation(wait_until="domcontentloaded", timeout=15000):
            await frame.locator('input[name="btnLogin"]').click()

        final_url = page.url
        print(f"Playwright final URL: {final_url}")
        cookies = await context.cookies()

        ok = "Monitor.aspx" in final_url or "p=" in final_url
        if ok:
            _gps_cookies = {c["name"]: c["value"] for c in cookies}
            print(f"Got cookies: {list(_gps_cookies.keys())}")
        return ok
    except Exception as e:
        print(f"Playwright login error: {e}")
        return False
    finally:
        try:
            if browser:
                await browser.close()
        except Exception:
            pass
        try:
            if pw:
                await pw.stop()
        except Exception:
            pass


async def _close_gps_session():
    global _gps_session
    sess, _gps_session = _gps_session, None
    if sess and not sess.closed:
        try:
            await sess.close()
        except Exception:
            pass


async def _invalidate_gps_session(sess: ClientSession | None = None):
    """Drop the cached cookies/session so the next call re-logs in.

    When `sess` is given, only invalidate if it is still the live session — a
    concurrent caller may already have rebuilt it, and we must not stomp the
    fresh one. The stale session is closed either way.
    """
    global _gps_cookies
    if sess is not None and sess is not _gps_session:
        if not sess.closed:
            try:
                await sess.close()
            except Exception:
                pass
        return
    _gps_cookies = None
    await _close_gps_session()


def _build_gps_session() -> ClientSession | None:
    cookies = _gps_cookies
    if not cookies:
        return None
    jar = CookieJar(unsafe=True)
    sess = ClientSession(cookie_jar=jar, timeout=GPS_HTTP_TIMEOUT)
    jar.update_cookies(cookies, response_url=URL(PLANET_GPS_BASE))
    return sess


async def ensure_gps_session() -> ClientSession | None:
    """Return a logged-in ClientSession, re-logging in when needed.

    Single-flight: concurrent callers share one login attempt instead of each
    launching a browser. Failed logins back off exponentially so a flood of
    requests can't turn into a browser storm.
    """
    global _gps_session, _login_fail_streak, _next_login_allowed

    if _gps_cookies and _gps_session and not _gps_session.closed:
        return _gps_session

    async with _gps_lock:
        if _gps_cookies and _gps_session and not _gps_session.closed:
            return _gps_session

        if not _gps_cookies:
            if time.monotonic() < _next_login_allowed:
                return None
            ok = await planet_gps_login_playwright()
            if not ok:
                _login_fail_streak += 1
                backoff = min(GPS_LOGIN_MAX_BACKOFF, 60 * 2 ** min(_login_fail_streak - 1, 5))
                _next_login_allowed = time.monotonic() + backoff
                print(f"GPS login failed, next attempt in {backoff}s")
                return None
            _login_fail_streak = 0
            _next_login_allowed = 0.0

        await _close_gps_session()
        _gps_session = _build_gps_session()
        return _gps_session


async def gps_post(path: str, payload: str, referer: str) -> dict | None:
    """POST to a PlanetGPS Ajax endpoint, re-logging in once if the session
    has expired. Returns the parsed JSON dict or None when GPS is unreachable.
    """
    headers = {**GPS_BASE_HEADERS, "Referer": referer}
    for attempt in range(2):
        sess = await ensure_gps_session()
        if sess is None:
            return None
        try:
            async with sess.post(f"{PLANET_GPS_BASE}{path}",
                                 data=payload, headers=headers) as resp:
                data = await resp.json(content_type=None)
            # A live PlanetGPS session always answers with a non-empty "d".
            # Missing/empty "d" means the cookie died — re-login and retry,
            # and never hand a bodyless response back to a handler as success.
            if isinstance(data, dict) and data.get("d"):
                return data
            print(f"GPS response missing data (attempt {attempt + 1}): {str(data)[:200]}")
        except Exception as e:
            print(f"GPS request error (attempt {attempt + 1}): {e}")
        await _invalidate_gps_session(sess)
    return None


def _parse_relaxed_json(raw: str) -> dict:
    """PlanetGPS returns JS object literals (unquoted keys, single quotes)."""
    fixed = re.sub(r'([{,])\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*:', r'\1"\2":', raw)
    fixed = re.sub(r":\s*'([^']*)'", r':"\1"', fixed)
    return json.loads(fixed)


# ================== MILEAGE HELPER ==================
# PlanetGPS truncates report periods longer than ~1 month, so long ranges are
# summed over ≤1-month windows.
GPS_MAX_WINDOW_DAYS = 28
GPS_MAX_WINDOWS = 24  # safety cap on the number of chunk requests (~up to ~2 years)


async def _window_mileage_km(device_id: str, start: str, end: str) -> float | None:
    """One PlanetGPS report call for a single window (≤ ~1 month).

    Returns distance in KM (float); 0.0 when the window has no data; or None
    when GPS is unreachable / unparseable, so the caller can decide to fail
    rather than report an undercount. start/end format: 'YYYY-MM-DD HH:MM'.
    """
    payload = json.dumps({
        "UserID": PLANET_GPS_USER_ID,
        "TimeZones": "5:00",
        "StartDates": start,
        "EndDates": end,
        "DeviceID": 0
    })
    data = await gps_post("/Ajax/ReportAjax.asmx/GetReportOverview", payload,
                          referer=f"{PLANET_GPS_BASE}/Report/Report.aspx")
    if not data or not data.get("d"):
        return None
    try:
        parsed = _parse_relaxed_json(data["d"])
    except Exception as e:
        print(f"Mileage parse error: {e}")
        return None

    rows = parsed.get("reports") or parsed.get("reportList") or parsed.get("devices") or parsed.get("list") or []
    if not rows:
        return 0.0

    row = next((r for r in rows if str(r.get("deviceID") or r.get("DeviceID") or "") == str(device_id)), None)
    if not row:
        row = await _match_row_by_device_name(rows, device_id)
    if not row:
        row = rows[0]

    raw = row.get("distance") or row.get("mileage") or row.get("Mileage")
    if raw in (None, "", "—"):
        return 0.0
    try:
        return float(raw)
    except Exception:
        return 0.0


async def get_period_mileage(device_id: str, start: str, end: str) -> str | None:
    """Mileage (in miles) for [start, end], summed over ≤1-month windows.

    PlanetGPS caps report periods at about a month, so a range longer than that
    (e.g. months since the last service) is split into GPS_MAX_WINDOW_DAYS-day
    chunks and summed. Returns a string number (miles) or None when GPS is
    unreachable. start/end format: 'YYYY-MM-DD HH:MM'.
    """
    try:
        start_dt = datetime.strptime(start, "%Y-%m-%d %H:%M")
        end_dt = datetime.strptime(end, "%Y-%m-%d %H:%M")
    except ValueError:
        return None
    if end_dt <= start_dt:
        return "0"

    total_km = 0.0
    got_any = False
    cur = start_dt
    windows = 0
    step = timedelta(days=GPS_MAX_WINDOW_DAYS)

    while cur < end_dt and windows < GPS_MAX_WINDOWS:
        w_end = min(cur + step, end_dt)
        km = await _window_mileage_km(
            device_id,
            cur.strftime("%Y-%m-%d %H:%M"),
            w_end.strftime("%Y-%m-%d %H:%M"),
        )
        if km is None:
            # GPS недоступен на этом окне — не отдаём заниженный итог.
            return None
        total_km += km
        got_any = True
        cur = w_end
        windows += 1

    if not got_any:
        return None
    # PlanetGPS reports km; the app shows miles
    return str(round(total_km * 0.621371, 2))


async def get_monthly_mileage(device_id: str, since: str | None = None) -> str | None:
    """Fetch mileage for the current month for a device from PlanetGPS.

    Если задан `since` (YYYY-MM-DD) и он позже начала текущего месяца — считаем
    пробег с этой даты, а не с 1-го числа (обнуление при смене водителя посреди
    месяца: новый водитель не наследует пробег предыдущего). Со следующего
    месяца since автоматически оказывается раньше начала месяца → счёт снова
    с 1-го.
    """
    now = datetime.now(timezone.utc)
    month_start = datetime(now.year, now.month, 1)
    start_dt = month_start
    if since:
        try:
            s = datetime.strptime(str(since)[:10], "%Y-%m-%d")
            if s > month_start:
                start_dt = s
        except Exception:
            pass
    start = start_dt.strftime("%Y-%m-%d 00:00")
    end = f"{now.year}-{now.month:02d}-{now.day:02d} 23:59"
    return await get_period_mileage(device_id, start, end)


async def _match_row_by_device_name(rows: list, device_id: str):
    """Report rows may lack device ids; match through the fleet list by name."""
    try:
        data = await gps_post("/Ajax/DevicesAjax.asmx/GetDevicesByUserID", FLEET_PAYLOAD,
                              referer=f"{PLANET_GPS_BASE}/Monitor.aspx")
        if not data or not data.get("d"):
            return None
        devices = _parse_relaxed_json(data["d"]).get("devices", [])
        device = next((d for d in devices if str(d.get("id", "")) == str(device_id)), None)
        if not device:
            return None
        dname = device.get("name", "").lower()
        return next((r for r in rows if r.get("name", "").lower() == dname), None)
    except Exception as e:
        print(f"Fleet lookup error: {e}")
        return None


# ================== ODOMETER CALIBRATION ==================
# Пробег ведём по трекеру, но «правду» задаёт одометр: админ периодически
# вбивает показание с фото одометра — это ЯКОРЬ. Дальше
#     реальный пробег = якорь + (мили трекера с даты якоря) * коэффициент,
# где коэффициент = (реальные мили / мили трекера) по прошлым сверкам.
# При каждой новой сверке коэффициент пересчитывается, а накопленная
# погрешность обнуляется — она живёт только внутри одного промежутка.
#
# PlanetGPS собственного одометра не отдаёт и здесь не нужен: используется
# только GetReportOverview (пройденное расстояние за период) — тот же вызов,
# что и раньше. Абсолютное показание приходит исключительно с фото.

SERVICE_INTERVAL_MILES = 7000
GPS_FACTOR_MIN = 0.70          # коэффициент вне этих границ считаем мусором
GPS_FACTOR_MAX = 1.30          # (опечатка в одометре, смена трекера и т.п.)
GPS_FACTOR_WINDOW = 5          # сколько последних сверок усредняем
GPS_FACTOR_MIN_SAMPLE = 50.0   # мили: на коротком плече коэффициент недостоверен


def _parse_day(value) -> datetime | None:
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d")
    except Exception:
        return None


def _to_float(value) -> float | None:
    """'128,970' / '128970 mi' / 128970.5 -> float. Иначе None."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    s = str(value).strip().replace(",", "").replace(" ", "")
    if not s:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def get_odometer_checks(driver: dict) -> list[dict]:
    """Валидные сверки одометра, отсортированные по дате (старые -> новые)."""
    out = []
    for c in driver.get("odometer_checks") or []:
        if not isinstance(c, dict):
            continue
        d = _parse_day(c.get("date"))
        m = _to_float(c.get("mileage"))
        if d is None or m is None:
            continue
        item = dict(c)
        item["mileage"] = m
        item["_date"] = d
        out.append(item)
    out.sort(key=lambda c: c["_date"])
    return out


def get_anchor(driver: dict) -> tuple[str, float, str] | None:
    """Точка отсчёта: (дата 'YYYY-MM-DD', показание одометра, источник).

    Кандидатов два: последняя сверка по фото одометра и пара
    last_service_mileage / last_service_date (её админ вбивает при ТО —
    это тоже реальное показание одометра). Берём более свежую: иначе,
    если ТО провели после последней сверки, расчёт уедет назад.
    Нет ни одного кандидата -> None, считать не от чего.
    """
    candidates = []
    checks = get_odometer_checks(driver)
    if checks:
        candidates.append((checks[-1]["_date"], checks[-1]["mileage"], "check"))
    d = _parse_day(driver.get("last_service_date"))
    m = _to_float(driver.get("last_service_mileage"))
    if d is not None and m is not None:
        candidates.append((d, m, "service"))
    if not candidates:
        return None
    # при равных датах выигрывает большее показание одометра
    best = max(candidates, key=lambda c: (c[0], c[1]))
    return best[0].strftime("%Y-%m-%d"), best[1], best[2]


def get_gps_factor(driver: dict) -> float:
    """Медиана коэффициентов по последним GPS_FACTOR_WINDOW сверкам.

    Медиана, а не среднее: одна кривая сверка (опечатка, снятый трекер)
    не должна утащить за собой расчёт. Нет данных -> 1.0 (без коррекции).
    """
    factors = []
    for c in get_odometer_checks(driver)[-GPS_FACTOR_WINDOW:]:
        f = _to_float(c.get("factor"))
        if f is not None and GPS_FACTOR_MIN <= f <= GPS_FACTOR_MAX:
            factors.append(f)
    if not factors:
        return 1.0
    factors.sort()
    mid = len(factors) // 2
    median = factors[mid] if len(factors) % 2 else (factors[mid - 1] + factors[mid]) / 2
    return round(median, 4)


async def gps_miles_between(device_id: str, start_day: str,
                            end_day: str | None = None) -> float | None:
    """Мили по трекеру за [start_day 00:00, end_day 23:59]. None = GPS недоступен."""
    if not device_id:
        return None
    d = _parse_day(start_day)
    if d is None:
        return None
    end_d = _parse_day(end_day) if end_day else None
    end = (end_d or datetime.now(timezone.utc)).strftime("%Y-%m-%d 23:59")
    raw = await get_period_mileage(str(device_id), d.strftime("%Y-%m-%d 00:00"), end)
    return _to_float(raw)


async def get_odometer_state(driver: dict) -> dict:
    """Текущий расчётный одометр и остаток до ТО.

    odometer = якорь + мили трекера с даты якоря * коэффициент.
    Все поля могут быть None — GPS отваливается регулярно, и врать
    заниженной цифрой хуже, чем честно показать «нет данных».
    """
    device_id = str(driver.get("planet_gps_device_id") or "").strip()
    checks = get_odometer_checks(driver)
    factor = get_gps_factor(driver)
    anchor = get_anchor(driver)

    state = {
        "odometer": None,
        "anchor": None,
        "anchor_source": anchor[2] if anchor else None,
        "gps_since_anchor": None,
        "gps_factor": factor,
        "checks_count": len(checks),
        "last_check_date": checks[-1]["_date"].strftime("%Y-%m-%d") if checks else None,
        "service_interval": SERVICE_INTERVAL_MILES,
        "service_base": _to_float(driver.get("last_service_mileage")),
        "miles_since_service": None,
        "miles_to_service": None,
        "has_gps": bool(device_id),
    }
    if not anchor:
        return state

    anchor_date, anchor_miles, _ = anchor
    state["anchor"] = {"date": anchor_date, "mileage": anchor_miles}

    gps = await gps_miles_between(device_id, anchor_date)
    if gps is None:
        return state

    state["gps_since_anchor"] = round(gps, 1)
    odo = anchor_miles + gps * factor
    state["odometer"] = round(odo, 1)

    base = state["service_base"]
    if base is not None:
        since = odo - base
        state["miles_since_service"] = round(since, 1)
        state["miles_to_service"] = round(SERVICE_INTERVAL_MILES - since, 1)
    return state


def find_driver_by_device(device_id: str) -> dict:
    """Водитель по id трекера — чтобы применить его коэффициент в /api/mileage."""
    did = str(device_id or "").strip()
    if not did:
        return {}
    for d in load_drivers().values():
        if str(d.get("planet_gps_device_id") or "").strip() == did:
            return d
    return {}


# ================== ROUTES ==================
async def handle_me(request: web.Request):
    init_data = request.headers.get("X-Telegram-Init-Data", "")
    user = verify_telegram_init_data(init_data) if init_data else None
    if not user:
        return web.json_response({"error": "Unauthorized"}, status=401)
    uid = user.get("id")
    # Имя берём из drivers.json (настоящее ФИО), а не из Telegram-профиля,
    # который может быть пустым/псевдонимом. На Telegram-имя откатываемся,
    # только если водителя нет в базе.
    driver = load_drivers().get(str(uid), {})
    name = (driver.get("name") or "").strip() or user.get("first_name", "")
    return web.json_response({
        "id": uid,
        "name": name,
        "is_admin": uid in ALLOWED_ADMINS,
        "is_driver": bool(driver),
    })


async def handle_drivers(request: web.Request):
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    drivers = load_drivers()
    items = list(drivers.items())
    file_lists = await asyncio.gather(*(
        asyncio.to_thread(list_driver_files, get_driver_folder(uid, d.get("name", "")))
        for uid, d in items
    ))
    result = []
    for (uid, d), files in zip(items, file_lists):
        result.append({
            "id": uid,
            "name": d.get("name", ""),
            "car_model": d.get("car_model", ""),
            "car_number": d.get("car_number", ""),
            "tariff": d.get("tariff", ""),
            "planet_gps_device_id": d.get("planet_gps_device_id", ""),
            "monthly_mileage_limit": d.get("monthly_mileage_limit", ""),
            "file_count": len(files),
            "gps_factor": get_gps_factor(d),
            "odometer_checks_count": len(get_odometer_checks(d)),
        })
    return web.json_response(result)


async def handle_driver_get(request: web.Request):
    """Admin view of one driver: profile + documents + this month's mileage.

    Returns the same shape the driver sees in their own profile (driver +
    files + mileage), so the admin panel can render the driver's profile
    'as the driver sees it' while staying authenticated as the admin.
    """
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    driver_id = request.match_info["id"]
    drivers = load_drivers()
    if driver_id not in drivers:
        return web.json_response({"error": "Driver not found"}, status=404)

    driver = drivers[driver_id]
    folder = get_driver_folder(driver_id, driver.get("name", ""))
    files = await asyncio.to_thread(list_driver_files, folder)

    # Live mileage for the current month (same source the driver self-view uses)
    device_id = driver.get("planet_gps_device_id", "")
    # Месячный пробег подтягивается на клиенте с fleet-бэкенда (там рабочий GPS),
    # поэтому здесь не блокируемся на медленном/ненадёжном скрейпинге.
    mileage_value = None
    now = datetime.now(timezone.utc)

    return web.json_response({
        "driver": {
            "id": driver_id,
            "name": driver.get("name", ""),
            "car_model": driver.get("car_model", ""),
            "car_year": driver.get("car_year", ""),
            "car_trim": driver.get("car_trim", ""),
            "car_number": driver.get("car_number", ""),
            "tariff": driver.get("tariff", ""),
            "planet_gps_device_id": driver.get("planet_gps_device_id", ""),
            "monthly_mileage_limit": driver.get("monthly_mileage_limit", ""),
            "weekly_price": driver.get("weekly_price", ""),
            "dmv_inspection_date": driver.get("dmv_inspection_date", ""),
            "payment_weekday": driver.get("payment_weekday", ""),
            "last_service_mileage": driver.get("last_service_mileage", ""),
            "last_service_date": driver.get("last_service_date", ""),
            # состояние калибровки — считается по локальному JSON, GPS не дёргаем
            "gps_factor": get_gps_factor(driver),
            "odometer_checks_count": len(get_odometer_checks(driver)),
        },
        "files": files,
        "odometer_checks": [
            {k: v for k, v in c.items() if k != "_date"}
            for c in get_odometer_checks(driver)
        ],
        "mileage": {
            "value": mileage_value,
            "limit": driver.get("monthly_mileage_limit", ""),
            "month": now.strftime("%B %Y"),
            "has_gps": bool(device_id),
        },
    })


async def handle_driver_add(request: web.Request):
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    try:
        body = await request.json()
        driver_id = str(body.get("id", "")).strip()
        name = body.get("name", "").strip()
        car_model = body.get("car_model", "").strip()
        car_number = body.get("car_number", "").strip()
        tariff = body.get("tariff", "").strip()

        if not all([driver_id, name]):
            return web.json_response({"error": "ID и имя обязательны"}, status=400)
        if not driver_id.isdigit():
            return web.json_response({"error": "ID must be numeric"}, status=400)

        drivers = load_drivers()
        if driver_id in drivers:
            return web.json_response({"error": "Driver with this ID already exists"}, status=409)

        record = {
            "name": name,
            "car_model": car_model,
            "car_number": car_number,
            "tariff": tariff,
        }
        # optional fields (stored as trimmed strings; empty if not provided)
        for field in OPTIONAL_DRIVER_FIELDS:
            record[field] = str(body.get(field, "")).strip()
        # сверки одометра — список, а не строка; в OPTIONAL_DRIVER_FIELDS
        # ему не место (там всё приводится к str)
        record["odometer_checks"] = []

        drivers[driver_id] = record
        save_drivers(drivers)
        return web.json_response({"success": True, "id": driver_id})
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)


async def handle_driver_update(request: web.Request):
    """Update driver fields (admin only). Used to set planet_gps_device_id and
    the custom monthly mileage limit on existing drivers."""
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    driver_id = request.match_info["id"]
    drivers = load_drivers()
    if driver_id not in drivers:
        return web.json_response({"error": "Driver not found"}, status=404)
    try:
        body = await request.json()
        # odometer_checks намеренно НЕ в списке: историю сверок правят
        # только через /driver/{id}/odometer, иначе её затрут строкой.
        updatable = ["car_model", "car_number", "tariff"] + OPTIONAL_DRIVER_FIELDS
        for field in updatable:
            if field in body:
                drivers[driver_id][field] = str(body[field]).strip()
        save_drivers(drivers)
        return web.json_response({"success": True})
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)


def _delete_driver_assets(folder: str):
    cloudinary.api.delete_resources_by_prefix(folder + "/")
    try:
        cloudinary.api.delete_folder(folder)
    except Exception:
        pass


async def handle_driver_delete(request: web.Request):
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    driver_id = request.match_info["id"]
    drivers = load_drivers()
    if driver_id not in drivers:
        return web.json_response({"error": "Driver not found"}, status=404)

    driver = drivers[driver_id]
    folder = get_driver_folder(driver_id, driver.get("name", ""))
    try:
        await asyncio.to_thread(_delete_driver_assets, folder)
    except Exception as e:
        print(f"Warn: couldn't delete Cloudinary resources: {e}")

    del drivers[driver_id]
    save_drivers(drivers)
    return web.json_response({"success": True})


async def handle_upload(request: web.Request):
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    driver_id = request.match_info["id"]
    drivers = load_drivers()
    if driver_id not in drivers:
        return web.json_response({"error": "Driver not found"}, status=404)

    driver = drivers[driver_id]
    folder = get_driver_folder(driver_id, driver.get("name", ""))
    try:
        body = await request.json()
        image_b64 = body.get("image")
        doc_name = body.get("name", "document").strip().replace(" ", "_")
        if not image_b64:
            return web.json_response({"error": "No image provided"}, status=400)
        if "," in image_b64:
            image_b64 = image_b64.split(",", 1)[1]
        image_bytes = base64.b64decode(image_b64)
        timestamp = int(time.time())
        filename = f"{doc_name}_{timestamp}"
        result = await asyncio.to_thread(
            cloudinary.uploader.upload, image_bytes,
            folder=folder, public_id=filename, resource_type="image"
        )
        return web.json_response({"success": True, "name": filename, "url": result["secure_url"], "public_id": result["public_id"]})
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)


async def handle_file_delete(request: web.Request):
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    try:
        body = await request.json()
        public_id = body.get("public_id", "").strip()
        if not public_id:
            return web.json_response({"error": "public_id required"}, status=400)
        await asyncio.to_thread(cloudinary.uploader.destroy, public_id, resource_type="image")
        return web.json_response({"success": True})
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)


# ================== DRIVER SELF-VIEW ==================
async def handle_driver_me(request: web.Request):
    """
    Driver-facing endpoint. Returns own profile + documents + monthly mileage.
    Auth: any valid Telegram user whose ID exists in drivers.json.
    """
    user = get_user_from_request(request)
    if not user:
        return web.json_response({"error": "Unauthorized"}, status=401)

    user_id = str(user.get("id", ""))
    drivers = load_drivers()
    if user_id not in drivers:
        return web.json_response({"error": "Not registered"}, status=403)

    driver = drivers[user_id]
    folder = get_driver_folder(user_id, driver.get("name", ""))
    files = await asyncio.to_thread(list_driver_files, folder)

    # Месячный пробег подтягивается на клиенте с fleet-бэкенда (рабочий GPS).
    mileage = None
    device_id = driver.get("planet_gps_device_id", "")
    now = datetime.now(timezone.utc)
    month_label = now.strftime("%B %Y")

    return web.json_response({
        "driver": {
            "id": user_id,
            "name": driver.get("name", ""),
            "car_model": driver.get("car_model", ""),
            "car_year": driver.get("car_year", ""),
            "car_trim": driver.get("car_trim", ""),
            "car_number": driver.get("car_number", ""),
            "tariff": driver.get("tariff", ""),
            "weekly_price": driver.get("weekly_price", ""),
            "planet_gps_device_id": device_id,
            "dmv_inspection_date": driver.get("dmv_inspection_date", ""),
            "payment_weekday": driver.get("payment_weekday", ""),
            "last_service_mileage": driver.get("last_service_mileage", ""),
            "last_service_date": driver.get("last_service_date", ""),
            "gps_factor": get_gps_factor(driver),
            "odometer_checks_count": len(get_odometer_checks(driver)),
        },
        "files": files,
        "mileage": {
            "value": mileage,
            "limit": driver.get("monthly_mileage_limit", ""),
            "month": month_label,
            "has_gps": bool(device_id),
        }
    })


async def handle_driver_me_upload(request: web.Request):
    """Driver uploads a document into their OWN Cloudinary folder.

    Auth: any valid Telegram user whose ID exists in drivers.json (same as
    /driver/me). This is the self-service counterpart to the admin-only
    /driver/{id}/upload — a driver may add files to their own folder but to
    no one else's.
    """
    user = get_user_from_request(request)
    if not user:
        return web.json_response({"error": "Unauthorized"}, status=401)

    user_id = str(user.get("id", ""))
    drivers = load_drivers()
    if user_id not in drivers:
        return web.json_response({"error": "Not registered"}, status=403)

    driver = drivers[user_id]
    folder = get_driver_folder(user_id, driver.get("name", ""))
    try:
        body = await request.json()
        image_b64 = body.get("image")
        doc_name = (body.get("name", "document") or "document").strip().replace(" ", "_") or "document"
        if not image_b64:
            return web.json_response({"error": "No image provided"}, status=400)
        if "," in image_b64:
            image_b64 = image_b64.split(",", 1)[1]
        image_bytes = base64.b64decode(image_b64)
        filename = f"{doc_name}_{int(time.time())}"
        result = await asyncio.to_thread(
            cloudinary.uploader.upload, image_bytes,
            folder=folder, public_id=filename, resource_type="image"
        )
        return web.json_response({
            "success": True,
            "name": filename,
            "url": result["secure_url"],
            "public_id": result["public_id"],
        })
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)


# ================== MILEAGE ENDPOINTS ==================
async def handle_mileage(request: web.Request):
    """GET /api/mileage/{device_id}?since=YYYY-MM-DD — monthly mileage for a device.

    Если задан ?since (дата пересадки водителя) и она в текущем месяце — пробег
    считается с этой даты, иначе с 1-го числа месяца.

    Цифра трекера домножается на коэффициент погрешности водителя (см. сверки
    одометра), поэтому месячный лимит считается по скорректированному пробегу.
    Ключ "mileage" остаётся строкой — контракт для фронта не меняется, сырое
    значение доступно в "mileage_raw".

    Intentionally unauthenticated to preserve the original public contract.
    """
    device_id = request.match_info["device_id"]
    if not device_id:
        return web.json_response({"mileage": None})
    since = request.rel_url.query.get("since")
    raw = await get_monthly_mileage(device_id, since)
    if raw is None:
        return web.json_response({"mileage": None, "mileage_raw": None, "gps_factor": 1.0})
    factor = get_gps_factor(find_driver_by_device(device_id))
    value = _to_float(raw)
    corrected = raw if value is None else str(round(value * factor, 2))
    return web.json_response({
        "mileage": corrected,
        "mileage_raw": raw,
        "gps_factor": factor,
    })


async def handle_service_mileage(request: web.Request):
    """GET /api/service-mileage/{device_id}?since=YYYY-MM-DD

    Мили, пройденные с последнего сервиса. Раньше это был «сырой» GPS с даты ТО;
    теперь — расчёт от якоря: реальное показание одометра + скорректированный
    GPS. Если сверок нет и калиброваться не от чего, поведение остаётся прежним.

    Возвращает прежний ключ "mileage" (строка) плюс блок "state" с расчётным
    одометром, коэффициентом и остатком до ТО.
    """
    device_id = request.match_info["device_id"]
    since = (request.rel_url.query.get("since", "") or "").strip()

    driver = find_driver_by_device(device_id)
    if driver:
        state = await get_odometer_state(driver)
        if state.get("miles_since_service") is not None:
            anchor = state.get("anchor") or {}
            return web.json_response({
                "mileage": str(state["miles_since_service"]),
                "since": anchor.get("date") or since,
                "state": state,
            })

    # Фолбэк: старое поведение — «сырой» GPS с даты сервиса.
    if not device_id or not since:
        return web.json_response({"mileage": None})
    d = _parse_day(since)
    if d is None:
        return web.json_response({"mileage": None})
    start = d.strftime("%Y-%m-%d 00:00")
    end = datetime.now(timezone.utc).strftime("%Y-%m-%d 23:59")
    mileage = await get_period_mileage(device_id, start, end)
    return web.json_response({"mileage": mileage, "since": start})


# ================== ODOMETER ENDPOINTS ==================
async def handle_odometer_get(request: web.Request):
    """GET /driver/{id}/odometer — история сверок + текущий расчётный одометр."""
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    driver_id = request.match_info["id"]
    drivers = load_drivers()
    if driver_id not in drivers:
        return web.json_response({"error": "Driver not found"}, status=404)
    driver = drivers[driver_id]
    checks = [{k: v for k, v in c.items() if k != "_date"}
              for c in get_odometer_checks(driver)]
    state = await get_odometer_state(driver)
    return web.json_response({"checks": checks, "state": state})


async def handle_odometer_add(request: web.Request):
    """POST /driver/{id}/odometer — админ вбивает показание с фото одометра.

    body: {"mileage": 131400, "date": "2026-08-08", "public_id": "...",
           "is_service": false, "note": "", "force": false}

    Сервер сам считает, сколько за тот же период показал трекер, и выводит
    погрешность и коэффициент.
    """
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    driver_id = request.match_info["id"]
    drivers = load_drivers()
    if driver_id not in drivers:
        return web.json_response({"error": "Driver not found"}, status=404)
    driver = drivers[driver_id]

    try:
        body = await request.json()
    except Exception:
        return web.json_response({"error": "Invalid JSON"}, status=400)

    mileage = _to_float(body.get("mileage"))
    if mileage is None or mileage <= 0:
        return web.json_response({"error": "mileage должен быть положительным числом"}, status=400)

    day = str(body.get("date") or "")[:10].strip() or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    day_dt = _parse_day(day)
    if day_dt is None:
        return web.json_response({"error": "date должен быть в формате YYYY-MM-DD"}, status=400)
    if day_dt > datetime.now(timezone.utc).replace(tzinfo=None) + timedelta(days=1):
        return web.json_response({"error": "date не может быть в будущем"}, status=400)

    force = bool(body.get("force"))
    prev = get_anchor(driver)
    device_id = str(driver.get("planet_gps_device_id") or "").strip()

    gps_since_prev = None
    real_since_prev = None
    factor = None
    drift = None

    if prev:
        prev_date, prev_miles, _ = prev
        prev_dt = _parse_day(prev_date)
        if day_dt < prev_dt and not force:
            return web.json_response(
                {"error": f"Дата сверки раньше предыдущей ({prev_date}). "
                          f"Отправьте force=true, если это осознанное исправление."},
                status=400)
        if mileage + 0.5 < prev_miles and not force:
            return web.json_response(
                {"error": f"Одометр {mileage:.0f} меньше предыдущего показания "
                          f"{prev_miles:.0f} от {prev_date}. Проверьте цифру или "
                          f"отправьте force=true."},
                status=400)
        if day_dt >= prev_dt and mileage >= prev_miles:
            real_since_prev = mileage - prev_miles
            gps_since_prev = await gps_miles_between(device_id, prev_date, day)
            if gps_since_prev is not None:
                drift = round(real_since_prev - gps_since_prev, 1)
                if gps_since_prev >= GPS_FACTOR_MIN_SAMPLE:
                    f = real_since_prev / gps_since_prev
                    if GPS_FACTOR_MIN <= f <= GPS_FACTOR_MAX:
                        factor = round(f, 4)
                    else:
                        print(f"Odometer check {driver_id}: коэффициент {f:.3f} вне "
                              f"[{GPS_FACTOR_MIN}, {GPS_FACTOR_MAX}] — не сохраняю")

    check = {
        "date": day,
        "mileage": round(mileage, 1),
        "gps_since_prev": round(gps_since_prev, 1) if gps_since_prev is not None else None,
        "real_since_prev": round(real_since_prev, 1) if real_since_prev is not None else None,
        "drift": drift,
        "drift_pct": (round(drift / gps_since_prev * 100, 2)
                      if drift is not None and gps_since_prev else None),
        "factor": factor,
        "source": str(body.get("source") or "photo"),
        "public_id": str(body.get("public_id") or ""),
        "note": str(body.get("note") or ""),
        "added_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    }

    checks = list(driver.get("odometer_checks") or [])
    checks.append(check)
    driver["odometer_checks"] = checks

    # Сверка сделана на сервисе — двигаем базу ТО на этот же одометр.
    if body.get("is_service"):
        driver["last_service_mileage"] = str(round(mileage, 1))
        driver["last_service_date"] = day

    save_drivers(drivers)
    state = await get_odometer_state(driver)
    return web.json_response({"success": True, "check": check, "state": state})


async def handle_odometer_delete(request: web.Request):
    """DELETE /driver/{id}/odometer — удалить ошибочную сверку.

    body: {"date": "2026-08-08"} или {"index": 2} (индекс в списке по дате).
    Админ вбивает цифры руками, опечатки неизбежны — нужен откат.
    """
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    driver_id = request.match_info["id"]
    drivers = load_drivers()
    if driver_id not in drivers:
        return web.json_response({"error": "Driver not found"}, status=404)
    driver = drivers[driver_id]

    try:
        body = await request.json()
    except Exception:
        body = {}

    checks = get_odometer_checks(driver)
    if not checks:
        return web.json_response({"error": "Сверок нет"}, status=404)

    idx = body.get("index")
    if idx is None:
        day = str(body.get("date") or "")[:10]
        matches = [i for i, c in enumerate(checks) if c.get("date") == day]
        if not matches:
            return web.json_response({"error": "Сверка с такой датой не найдена"}, status=404)
        idx = matches[-1]
    try:
        idx = int(idx)
        removed = checks.pop(idx)
    except (ValueError, TypeError, IndexError):
        return web.json_response({"error": "Некорректный index"}, status=400)

    driver["odometer_checks"] = [{k: v for k, v in c.items() if k != "_date"} for c in checks]
    save_drivers(drivers)
    state = await get_odometer_state(driver)
    return web.json_response({
        "success": True,
        "removed": {k: v for k, v in removed.items() if k != "_date"},
        "state": state,
    })


async def handle_odometer_me(request: web.Request):
    """GET /api/odometer/me — водитель видит свой расчётный пробег и остаток до ТО."""
    user = get_user_from_request(request)
    if not user:
        return web.json_response({"error": "Unauthorized"}, status=401)
    user_id = str(user.get("id", ""))
    drivers = load_drivers()
    if user_id not in drivers:
        return web.json_response({"error": "Not registered"}, status=403)
    state = await get_odometer_state(drivers[user_id])
    return web.json_response(state)


# ================== FLEET ==================
async def handle_fleet(request: web.Request):
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    data = await gps_post("/Ajax/DevicesAjax.asmx/GetDevicesByUserID", FLEET_PAYLOAD,
                          referer=f"{PLANET_GPS_BASE}/Monitor.aspx")
    print(f"Fleet response: {str(data)[:200]}")
    if not data or not data.get("d"):
        return web.json_response({"error": "PlanetGPS unavailable"}, status=502)
    return web.json_response(data)


async def handle_fleet_report(request: web.Request):
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    params = request.rel_url.query
    start = params.get("from", "")
    end = params.get("to", "")
    if not start or not end:
        return web.json_response({"error": "Missing from/to"}, status=400)

    payload = json.dumps({
        "UserID": PLANET_GPS_USER_ID,
        "TimeZones": "5:00",
        "StartDates": start,
        "EndDates": end,
        "DeviceID": 0
    })
    data = await gps_post("/Ajax/ReportAjax.asmx/GetReportOverview", payload,
                          referer=f"{PLANET_GPS_BASE}/Report/Report.aspx?id={PLANET_GPS_USER_ID}&deviceid=0&randon=21896")
    print(f"Report response: {str(data)[:300]}")
    if not data:
        return web.json_response({"error": "PlanetGPS unavailable"}, status=502)
    return web.json_response(data)


async def _fetch_report_page(report_url: str, headers: dict) -> str | None:
    """GET the report page HTML, re-logging in once if the session is dead."""
    for attempt in range(2):
        sess = await ensure_gps_session()
        if sess is None:
            return None
        try:
            async with sess.get(report_url, headers=headers,
                                timeout=GPS_REPORT_TIMEOUT) as resp:
                html = await resp.text()
            if 'id="__VIEWSTATE"' in html:
                return html
            print(f"Report page has no viewstate (attempt {attempt + 1})")
        except Exception as e:
            print(f"Report page fetch error (attempt {attempt + 1}): {e}")
        await _invalidate_gps_session(sess)
    return None


def _convert_report_to_miles(content: bytes, start: str, end: str) -> tuple[bytes, str, str]:
    """Convert the km mileage column of the exported report to miles.

    openpyxl is imported lazily so a missing optional dependency can only
    degrade this one feature (the report is returned unconverted) instead of
    crashing the whole bot+API at startup.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        for ws in wb.worksheets:
            header_row = None
            mileage_col = None
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        v = cell.value.strip().lower()
                        if v in ('mileage', 'distance', 'пробег', 'miles'):
                            mileage_col = cell.column
                            header_row = cell.row
                if header_row:
                    break
            if mileage_col and header_row:
                for row in ws.iter_rows(min_row=header_row + 1):
                    for cell in row:
                        if cell.column == mileage_col and cell.value is not None:
                            try:
                                cell.value = round(float(cell.value) * 0.621371, 2)
                            except Exception:
                                pass
                header_cell = ws.cell(row=header_row, column=mileage_col)
                if header_cell.value:
                    header_cell.value = "Mileage (mi)"
        out = io.BytesIO()
        wb.save(out)
        return (
            out.getvalue(),
            f"report_{start[:10]}_{end[:10]}_miles.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        print(f"Excel conversion error: {e}")
        return (
            content,
            f"report_{start[:10]}_{end[:10]}.xls",
            "application/vnd.ms-excel",
        )


async def handle_fleet_report_excel(request: web.Request):
    if not is_admin_request(request):
        return web.json_response({"error": "Forbidden"}, status=403)
    params = request.rel_url.query
    start = params.get("from", "")
    end = params.get("to", "")
    if not start or not end:
        return web.json_response({"error": "Missing from/to"}, status=400)

    report_url = f"{PLANET_GPS_BASE}/Report/Report.aspx?id={PLANET_GPS_USER_ID}&deviceid=0&randon=12345"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": report_url,
    }

    html = await _fetch_report_page(report_url, headers)
    if html is None:
        return web.json_response({"error": "PlanetGPS unavailable"}, status=502)

    sess = _gps_session
    if sess is None or sess.closed:
        return web.json_response({"error": "PlanetGPS unavailable"}, status=502)

    fields = {}
    for field in ["__VIEWSTATE", "__VIEWSTATEGENERATOR", "__EVENTVALIDATION"]:
        m = re.search(rf'id="{field}"[^>]*value="([^"]*)"', html)
        fields[field] = m.group(1) if m else ""

    data = {
        **fields,
        "beginTime": start,
        "endTime": end,
        "btnToExcel": "To Excel",
        "hidUserID": str(PLANET_GPS_USER_ID),
        "hidTimeZone": "5:00",
        "hidDeviceID": "0",
    }
    post_headers = {
        **headers,
        "Content-Type": "application/x-www-form-urlencoded",
    }
    try:
        async with sess.post(report_url, data=data, headers=post_headers,
                             timeout=GPS_REPORT_TIMEOUT) as resp:
            content = await resp.read()
        content, filename, content_type = await asyncio.to_thread(
            _convert_report_to_miles, content, start, end
        )
        return web.Response(
            body=content,
            content_type=content_type,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)


# ================== CORS MIDDLEWARE ==================
@web.middleware
async def cors_middleware(request, handler):
    if request.method == "OPTIONS":
        response = web.Response()
    else:
        try:
            response = await handler(request)
        except web.HTTPException as ex:
            response = ex
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Telegram-Init-Data"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PATCH, DELETE, OPTIONS"
    return response


def create_app() -> web.Application:
    app = web.Application(middlewares=[cors_middleware])
    app.router.add_get("/me", handle_me)
    app.router.add_get("/drivers", handle_drivers)
    app.router.add_post("/drivers", handle_driver_add)
    app.router.add_get("/driver/me", handle_driver_me)
    app.router.add_post("/driver/me/upload", handle_driver_me_upload)
    app.router.add_get("/driver/{id}", handle_driver_get)
    app.router.add_patch("/driver/{id}", handle_driver_update)
    app.router.add_delete("/driver/{id}", handle_driver_delete)
    app.router.add_post("/driver/{id}/upload", handle_upload)
    app.router.add_get("/driver/{id}/odometer", handle_odometer_get)
    app.router.add_post("/driver/{id}/odometer", handle_odometer_add)
    app.router.add_delete("/driver/{id}/odometer", handle_odometer_delete)
    app.router.add_delete("/file", handle_file_delete)
    app.router.add_get("/api/fleet", handle_fleet)
    app.router.add_get("/api/odometer/me", handle_odometer_me)
    app.router.add_get("/api/mileage/{device_id}", handle_mileage)
    app.router.add_get("/api/service-mileage/{device_id}", handle_service_mileage)
    app.router.add_get("/api/fleet/report", handle_fleet_report)
    app.router.add_get("/api/fleet/report/excel", handle_fleet_report_excel)

    options_paths = [
        "/me", "/drivers", "/driver/me", "/driver/me/upload", "/driver/{id}",
        "/driver/{id}/upload", "/driver/{id}/odometer", "/file", "/api/fleet",
        "/api/odometer/me",
        "/api/mileage/{device_id}",
        "/api/service-mileage/{device_id}",
        "/api/fleet/report", "/api/fleet/report/excel"
    ]
    for path in options_paths:
        app.router.add_options(path, lambda r: web.Response())
    return app


async def _gps_ping() -> bool:
    """Cheap HTTP request that both checks and extends the server-side session."""
    sess = _gps_session
    if not _gps_cookies or sess is None or sess.closed:
        return False
    try:
        headers = {**GPS_BASE_HEADERS, "Referer": f"{PLANET_GPS_BASE}/Monitor.aspx"}
        async with sess.post(f"{PLANET_GPS_BASE}/Ajax/DevicesAjax.asmx/GetDevicesByUserID",
                             data=FLEET_PAYLOAD, headers=headers) as resp:
            data = await resp.json(content_type=None)
        return isinstance(data, dict) and bool(data.get("d"))
    except Exception as e:
        print(f"GPS ping error: {e}")
        return False


async def keep_gps_session_alive():
    """Keep the PlanetGPS session warm with cheap HTTP pings; only do a full
    browser login when the session is actually gone. (The old version
    launched a fresh browser every cycle and never closed the previous one —
    the leaked processes eventually exhausted the container's threads and
    every subsequent launch failed.)
    """
    global _next_login_allowed
    await asyncio.sleep(30)
    while True:
        try:
            if not await _gps_ping():
                print("Background GPS session refresh...")
                # The keepalive cadence itself rate-limits logins, so clear the
                # request-driven backoff here: a transient failure must never
                # leave the session frozen until the (up to 30-min) backoff ends.
                _next_login_allowed = 0.0
                await _invalidate_gps_session()
                if await ensure_gps_session():
                    print("GPS session refreshed OK")
                else:
                    print("GPS session refresh FAILED")
        except Exception as e:
            print(f"GPS keepalive error: {e}")
        await asyncio.sleep(GPS_PING_INTERVAL)


async def start_server():
    app = create_app()
    runner = web.AppRunner(app)
    await runner.setup()
    port = int(os.getenv("PORT", 8080))
    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()
    print(f"API server running on port {port}")

    keepalive = asyncio.create_task(keep_gps_session_alive())
    try:
        await asyncio.Event().wait()
    finally:
        keepalive.cancel()
        await _close_gps_session()
        await runner.cleanup()


if __name__ == "__main__":
    asyncio.run(start_server())
